from openpyxl import load_workbook#导入所需包
file_excel = load_workbook("数据.xlsx")#读文件
sheetnames = file_excel.get_sheet_names()#获取读文件中所有的sheet
ws = file_excel.get_sheet_by_name(sheetnames[0])#获取第一个sheet内容
rows_max = ws.max_row#获取sheet的最大行数
cols_max = ws.max_column#获取sheet的最大列数
print('行数',str(rows_max))
print('列数',str(cols_max))
for r in range(1,rows_max+1):
    for c in range(1,cols_max+1):
        print(ws.cell(r,c).value)
    if r==5: #显示前五行
        break
saveExcel = "数据.xlsx"
file_excel.save(saveExcel)  
from openpyxl import Workbook

test1 = Workbook()  # 打开一个将写的文件
sheet1 = test1.create_sheet(index=0)  # 在将写的文件创建sheet
row1 = ['姓名','数学','语文']
row2 = ['tom',78,89]
sheet1.append(row1)#使用append插入数据
sheet1.append(row2)
sheet1['A3'] = '小明'#直接插入数据
sheet1['B3'] = 80
sheet1['C3'] = 84

saveExcel = "test1.xlsx"
test1.save(saveExcel)  # 保存


